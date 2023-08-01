import pandas as pd
import pyodbc
import json
from django.shortcuts import render, redirect
from .models import Societe, AssociationSociete, Type, Tiers, Indentite, IntercoHistorique, Tableau
from django.http import JsonResponse, HttpResponseBadRequest
from django.core.exceptions import ValidationError
from .forms import ImportExcelForm
from openpyxl import load_workbook
from django.contrib import messages


def get_connection_string(database):
    # Utilisation de f-string pour formater la chaîne de connexion
    value = Societe.objects.get(name=database)

    # Définition des serveurs à essayer dans l'ordre de priorité
    server = value.server
    username = value.user
    password = value.password
    error_message = None  # Ajoutez une variable pour capturer l'erreur
    try:
        conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};" \
                   f"DATABASE={database};UID={username};PWD={password}"
        print(f"CONNEXION STR {conn_str}")
        # Tentative de connexion au serveur actuel
        with pyodbc.connect(conn_str) as connection:
            # Test de la connexion au serveur en exécutant une requête d'exemple
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
            print(f'Connexion réussie au serveur : {server}')
            print(f"Output {conn_str}")
            return [conn_str, server]
    except pyodbc.OperationalError as e:
        error_message = f'Échec de la connexion au serveur : {server}, Erreur : {e}'

    # Si aucun des serveurs n'a réussi à se connecter, on lève une exception
    if error_message:
        raise ValidationError(error_message)


def chercher(database, ct_num, cg_num, conn_str):
    # Utilisation de la fonction get_connection_string pour obtenir la chaîne de connexion au serveur approprié
    # conn_str = get_connection_string(database)
    # Connexion à la base de données
    error_message = None  # Ajoutez une variable pour capturer l'erreur
    try:
        with pyodbc.connect(conn_str[0]) as connection:
            # Création d'un curseur
            cursor = connection.cursor()

            ct_num = format_tuple(ct_num)
            cg_num = format_tuple(cg_num)

            # Utilisation de paramètres pour éviter les vulnérabilités d'injection SQL
            query = "SELECT  JO_NUM, EC_PIECE, CG_NUM, CT_NUM, EC_INTITULE, EC_REFPIECE, EC_No, CASE WHEN EC_SENS =0 THEN EC_MONTANT END AS DEBIT, " \
                    f" CASE WHEN EC_SENS =1 THEN EC_MONTANT END AS CREDIT, LETTRE_INTERCO FROM F_ECRITUREC WHERE (CT_NUM in {ct_num}" \
                    f" OR CG_NUM in {cg_num}) AND year(jm_date)=2023"

            cursor.execute(query)

            # Récupération des résultats
            rows = cursor.fetchall()

    except pyodbc.OperationalError as e:
        error_message = f"Une erreur s'est produite lors de la connexion à la base de données INVISO : {e}"

    if error_message:
        raise ValidationError(error_message)

    # Liste pour stocker les résultats
    data = []

    for row in rows:
        JO_NUM, EC_PIECE, CG_NUM, CT_NUM, EC_INTITULE, EC_REFPIECE, EC_No, DEBIT, CREDIT, LETTRE_INTERCO = row

        item = {
            'jo_num': JO_NUM,
            'ec_piece': EC_PIECE,
            'ec_no': EC_No,
            'cg_num': CG_NUM,
            'ct_num': CT_NUM,
            'ec_intitule': EC_INTITULE,
            'ec_refpiece': EC_REFPIECE,
            'interco': LETTRE_INTERCO,
            'debit': "{:.2f}".format(float(DEBIT)) if DEBIT else '',
            'credit': "{:.2f}".format(float(CREDIT)) if CREDIT else ''
        }
        data.append(item)
    return data


def updade(database, ecPiece, ecRefPiece, ecNo, conn_str, jo_num, interco):
    error_message = None
    try:
        with pyodbc.connect(conn_str[0]) as connection:
            cursor = connection.cursor()
            query = "UPDATE F_ECRITUREC SET LETTRE_INTERCO=? WHERE EC_PIECE=? AND EC_REFPIECE=? AND EC_No=? AND JO_NUM=? AND year(jm_date)=2023"
            params = (interco, ecPiece, ecRefPiece, ecNo, jo_num)
            cursor.execute(query, params)

            # Just return True after the successful update

            return True
    except pyodbc.OperationalError as e:
        error_message = f"Une erreur s'est produite lors de la connexion à la base de données INVISO : {e}"
    except pyodbc.Error as e:
        error_message = f"Une erreur s'est produite lors de l'exécution de la requête : {e}"
        return False
    finally:
        if error_message:
            # You may log the error or take appropriate action based on the error_message.
            pass


def format_tuple(t):
    if len(t) == 1:
        return "('{}')".format(t[0])  # Ajoute des guillemets simples et supprime la virgule
    return "({})".format(', '.join("'{}'".format(value) for value in t))


def inter_value(value):
    ct_num = []  # Utiliser un tuple vide au lieu d'une chaîne de caractères
    cg_num = []  # Utiliser un tuple vide au lieu d'une chaîne de caractères
    for ass in value:
        if ass.type.intitule == 'CT':
            ct_num.append(ass.tiers.value)  # Utiliser une virgule pour créer un tuple avec une seule valeur
        else:
            cg_num.append(ass.tiers.value)  # Utiliser une virgule pour créer un tuple avec une seule valeur

    # Créer une liste de tuples avec les tuples contenant une seule valeur
    result = [tuple(ct_num), tuple(cg_num)]
    print(f'R : {result} ct_num : {ct_num} et cg_num: {cg_num}')
    return result


def index(request):
    results1 = []
    results2 = []
    message = None

    associations1 = {}
    associations2 = {}

    societe_1_name = ''
    societe_2_name = ''

    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['excel_file']
            try:
                importer_associations(file)
                messages.success(request, 'Importation réussie !')
                return redirect('home:index')
            except ValidationError as e:
                messages.error(request, f"Une erreur s'est produite lors du traitement du fichier excel : {str(e)}")
            except Exception as e:
                messages.error(request, f"Une erreur s'est produite lors du traitement du fichier excel : {str(e)}")
        else:
            messages.error(request, 'Formulaire invalide, veuillez vérifier les champs !')

    else:
        form = ImportExcelForm()
        societe_1_name = request.GET.get('societe_1_name', '')
        societe_2_name = request.GET.get('societe_2_name', '')

        if societe_1_name and societe_2_name and societe_1_name != societe_2_name:
            try:
                associations1 = AssociationSociete.objects.filter(
                    societe1__name__exact=societe_1_name,
                    societe2__name__exact=societe_2_name,
                )

                associations2 = AssociationSociete.objects.filter(
                    societe1__name__exact=societe_2_name,
                    societe2__name__exact=societe_1_name,
                )

                value1 = inter_value(associations1)
                value2 = inter_value(associations2)

                if not value1[0]:
                    value1[0] = ('',)
                if not value1[1]:
                    value1[1] = ('',)
                if not value2[0]:
                    value2[0] = ('',)
                if not value2[1]:
                    value2[1] = ('',)

                conn_1 = get_connection_string(societe_1_name)
                conn_2 = get_connection_string(societe_2_name)

                results1 = chercher(database=societe_1_name, ct_num=value1[0], cg_num=value1[1], conn_str=conn_1)
                results2 = chercher(database=societe_2_name, ct_num=value2[0], cg_num=value2[1], conn_str=conn_2)

            except ValidationError as e:
                message = f"Une erreur s'est produite lors de la recherche dans la base de données (ValidationError) :  {str(e)}"
            except Exception as e:
                message = f"Une erreur s'est produite lors de la recherche dans la base de données (Exception) :  {str(e)}"

        else:
            associations1 = {}
            associations2 = {}

    societe = Societe.objects.filter(active=True)

    context = {
        'associations1': associations1,
        'associations2': associations2,
        'results1': results1,
        'results2': results2,
        'societes': societe,
        'societe_1_name': societe_1_name,
        'societe_2_name': societe_2_name,
        'form': form,
        'message': message
    }
    return render(request, 'home/index.html', context)


def import_data(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['excel_file']
            try:
                importer_associations(file)
                success_message = 'Importation réussie !'
                return JsonResponse({'success_message': success_message})
            except ValidationError as e:
                error_message = str(e)
                return JsonResponse({'error_message': error_message})
    else:
        form = ImportExcelForm()
    return JsonResponse({'form': form})


def identite_check(value):
    value = str(value).strip()
    try:
        identite_obj = Indentite.objects.filter(name__exact=value).first()
        return identite_obj.correspondance if identite_obj else value
    except Indentite.DoesNotExist:
        return value


def importer_associations(file):
    wb = load_workbook(file)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        df = pd.read_excel(file, sheet_name=sheet_name)
        required_columns = ['SOCIETE1', 'TIERS', 'SOCIETE2', 'TYPE']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValidationError(
                f"Les colonnes suivantes sont manquantes dans la feuille '{sheet_name}': {', '.join(missing_columns)}")

        for index, row in df.iterrows():
            societe1_name = row['SOCIETE1']
            tiers_value = row['TIERS']
            societe2_name = row['SOCIETE2']
            type_intitule = row['TYPE']

            societe1, _ = Societe.objects.get_or_create(name=identite_check(societe1_name))
            societe2, _ = Societe.objects.get_or_create(name=identite_check(societe2_name))
            tiers, _ = Tiers.objects.get_or_create(value=str(tiers_value).strip())
            type_obj, _ = Type.objects.get_or_create(intitule=str(type_intitule).strip())

            AssociationSociete.objects.get_or_create(
                societe1=societe1,
                societe2=societe2,
                tiers=tiers,
                type=type_obj,
                # connexion=connexion
                # Associer la connexion à la société
            )


# Nouvelle fonction pour générer le numéro interco unique
# Nouvelle fonction pour générer le numéro interco unique
def generate_interco():
    last_interco_obj = IntercoHistorique.objects.order_by('-id').first()

    if last_interco_obj and last_interco_obj.interco.startswith('I'):
        last_number = int(last_interco_obj.interco[1:])  # Récupérer le numéro après "I"
        new_number = last_number + 1
    else:
        new_number = 1

    interco_number = f"{new_number:04}"
    return f"I{interco_number}"


# Nouvelle fonction pour mettre à jour les données de A et B dans la base de données
def update_data(database, ecPiece, ecRefPiece, ecNo, joNum, interco):
    update = updade(database=database, ecPiece=ecPiece, ecRefPiece=ecRefPiece, ecNo=ecNo,
                    conn_str=get_connection_string(database), jo_num=joNum, interco=interco)

    return update


# Refactorisation de la fonction ajax_view

def ajax_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            A = data.get('A', {})
            B = data.get('B', {})
            debit_A = A.get('debit', 0)
            credit_A = A.get('credit', 0)
            debit_B = B.get('debit', 0)
            credit_B = B.get('credit', 0)
            print("Dans post !")

            if debit_A == credit_B and credit_A == debit_B:
                print("A = B")
                interco = generate_interco()  # Générer le numéro d'interco

                IntercoHistorique.objects.create(
                    interco=interco,
                    tableau1=Tableau.objects.create(base=A['base'], ecPiece=A['ecPiece'],
                                                    ecRefPiece=A['ecRefPiece'], ecNo=A['ecNo']),
                    tableau2=Tableau.objects.create(base=B['base'], ecPiece=B['ecPiece'],
                                                    ecRefPiece=B['ecRefPiece'], ecNo=B['ecNo']),
                )

                value1 = update_data(database=A.get('base', ''), ecPiece=A.get('ecPiece', ''),
                                     ecRefPiece=A.get('ecRefPiece', ''),
                                     ecNo=A.get('ecNo', ''), joNum=A.get('joNum', ''), interco=interco)

                value2 = update_data(database=B.get('base', ''), ecPiece=B.get('ecPiece', ''),
                                     ecRefPiece=B.get('ecRefPiece', ''),
                                     ecNo=B.get('ecNo', ''), joNum=B.get('joNum', ''), interco=interco)

                if value1 and value2:
                    message = "Update reuissi !"
                    # return JsonResponse({'success': True, 'message': message, 'societe_1_name': A.get('base', ''),'societe_2_name': B.get('base', '')})
                    return redirect('home:index')
                else:
                    message = 'An error occurred during data update.'
                    return JsonResponse({'success': False, 'message': message})

            else:
                message = 'Debit/Credit or ecRefPiece values do not match.'
                return JsonResponse({'success': False, 'message': message})

        except json.JSONDecodeError:
            message = "Invalid JSON data."
            return JsonResponse({'success': False, 'message': message})

        except Exception as e:
            message = f'An unexpected error occurred : {e}'
            return JsonResponse({'success': False, 'message': message})
    else:
        message = "Invalid request method."
        return JsonResponse({'success': False, 'message': message})
