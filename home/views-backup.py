import pandas as pd
import pyodbc
from django.shortcuts import render, redirect
from .models import Societe, AssociationSociete, Type, Tiers, Connexion
from django.http import JsonResponse

from django.core.exceptions import ValidationError
from .forms import ImportExcelForm
from openpyxl import load_workbook

server = '192.168.1.161'
username = 'reader'
password = 'm1234'


def chercher(base):
    database = str(base.db_name)

    # Chaîne de connexion
    conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}"
    # Connexion à la base de données

    with pyodbc.connect(conn_str) as connection:
        # Création d'un curseur
        cursor = connection.cursor()

        # if database == 'FLY TECHNOLOGIES':
        #     value = ('C22013', 'F00057')
        #
        # elif database == 'IDMOTORS':
        #     value = ('C22272', 'F22725')
        #     value2 = ('C22272', 'F22725')

        # Exécution d'une requête SELECT
        query = "SELECT EC_PIECE, CG_NUM, EC_INTITULE, CASE WHEN EC_SENS =0 THEN EC_MONTANT END AS DEBIT," \
                f" CASE WHEN EC_SENS =1 THEN EC_MONTANT END AS CREDIT FROM F_ECRITUREC WHERE (CT_NUM in {value} OR CG_NUM in {value2}) " \
                "AND year(jm_date)=2023"

        cursor.execute(query)

        # Récupération des résultats
        rows = cursor.fetchall()

    # Liste pour stocker les résultats
    data = []

    for row in rows:
        EC_PIECE, CG_NUM, EC_INTITULE, DEBIT, CREDIT = row

        item = {
            'ec_piece': EC_PIECE,
            'cg_num': str(CG_NUM)[:5],
            'ec_intitule': EC_INTITULE,
            'debit': "{:.2f}".format(float(DEBIT)) if DEBIT else '',
            'credit': "{:.2f}".format(float(CREDIT)) if CREDIT else ''
        }

        data.append(item)

    return data


def inter_value(value):
    r = ()
    for ass in value:
        r += (ass.tiers,)
    return r


def index(request):
    if request.session.get('user_session', True):
        bases = Societe.objects.filter(active=True)
        if request.method == 'POST':
            form = ImportExcelForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['excel_file']
                try:
                    importer_associations(file)
                    message = 'Importation réussie !'
                    return redirect('home:index')
                except ValidationError as e:
                    message = str(e)
                    return redirect('home:index')

            context = {
                'form': form,
                'bases': bases
            }

            return render(request, 'home/index.html', context)
        else:
            form = ImportExcelForm()
            societe_1_name = request.GET.get('societe_1_name', '')
            societe_2_name = request.GET.get('societe_2_name', '')

            associations_normals = AssociationSociete.objects.filter(
                societe1__name__exact=societe_1_name,
                societe2__name__exact=societe_2_name,
            )

            associations_invers = AssociationSociete.objects.filter(
                societe1__name__exact=societe_2_name,
                societe2__name__exact=societe_1_name,
            )

            value_tiers = tuple(inter_value(associations_invers) + inter_value(associations_normals))
            if value_tiers is not None:
                chercher()

            societe = Societe.objects.filter(active=True)

            context = {
                'associations_normals': associations_normals,
                'associations_invers': associations_invers,
                'societes': societe,
                'societe_1_name': societe_1_name,
                'societe_2_name': societe_2_name,
                'form': form
            }
            return render(request, 'home/index.html', context)
    else:
        return redirect('auth:login')


def base_details(request, base_id):
    base = Societe.objects.select_related('societe').get(pk=base_id)

    lists = chercher(base)
    print("------------------------------------------------------------------")
    print(lists)
    print("------------------------------------------------------------------")
    data = {
        'lists': lists,
    }

    return JsonResponse(data)


def import_data(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['excel_file']
            try:
                importer_associations(file)
                success_message = 'Importation réussie !'
                return JsonResponse({'success_message': success_message})
            except ValidationError as e:
                error_message = str(e)
                return JsonResponse({'error_message': error_message})
    else:
        form = ImportExcelForm()
    return JsonResponse({'form': form})


def importer_associations(file):
    wb = load_workbook(file)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        df = pd.read_excel(file, sheet_name=sheet_name)
        required_columns = ['S1', 'TS', 'S2', 'TP']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValidationError(
                f"Les colonnes suivantes sont manquantes dans la feuille '{sheet_name}': {', '.join(missing_columns)}")

        for index, row in df.iterrows():
            societe1_name = row['S1']
            tiers_value = row['TS']
            societe2_name = row['S2']
            type_intitule = row['TP']

            societe1, _ = Societe.objects.get_or_create(name=societe1_name)
            societe2, _ = Societe.objects.get_or_create(name=societe2_name)
            tiers, _ = Tiers.objects.get_or_create(value=tiers_value)
            type_obj, _ = Type.objects.get_or_create(intitule=type_intitule)

            AssociationSociete.objects.create(
                societe1=societe1,
                societe2=societe2,
                tiers=tiers,
                type=type_obj,
                # connexion=connexion  # Associer la connexion à la société
            )
